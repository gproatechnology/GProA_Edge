import openpyxl
import os

uploads_dir = r'c:\Users\X1\OneDrive\Documentos\Python_VS Code\GProA\GProA_EOSIS_Edge\backend\uploads'
path = os.path.join(uploads_dir, '38b4a51d-ff33-4c9a-9216-090bc64828d6.xlsx')

print(f"Inspecting Excel: {os.path.basename(path)}")

try:
    wb = openpyxl.load_workbook(path, data_only=True)
    print(f"Sheets: {wb.sheetnames}")
    
    for sheet_name in wb.sheetnames:
        print(f"\n--- Sheet: {sheet_name} ---")
        ws = wb[sheet_name]
        # Mostrar primeras 10 filas y 5 columnas
        for row in ws.iter_rows(max_row=10, max_col=5):
            print([cell.value for cell in row])

except Exception as e:
    print(f"Error reading Excel: {e}")
