import openpyxl
from copy import copy

def copy_sheet(source_sheet, target_sheet):
    # Copy values and styles
    for row in source_sheet:
        for cell in row:
            new_cell = target_sheet.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.border = copy(cell.border)
                new_cell.fill = copy(cell.fill)
                new_cell.number_format = copy(cell.number_format)
                new_cell.protection = copy(cell.protection)
                new_cell.alignment = copy(cell.alignment)
    
    # Copy column widths
    for col_letter, col_dimension in source_sheet.column_dimensions.items():
        target_sheet.column_dimensions[col_letter].width = col_dimension.width
        
    # Copy row heights
    for row_num, row_dimension in source_sheet.row_dimensions.items():
        target_sheet.row_dimensions[row_num].height = row_dimension.height

    # Copy merged cells
    for merged_cell_range in source_sheet.merged_cells.ranges:
        target_sheet.merge_cells(str(merged_cell_range))

wbs_path = r"c:\Users\X1\OneDrive\Documentos\Python_VS Code\GProA\GProA_EOSIS_Edge\backend\app\assets\wbs_template.xlsx"
wwr_path = r"c:\Users\X1\OneDrive\Documentos\Python_VS Code\GProA\GProA_EOSIS_Edge\docs\Documentos_EOSIS\EEM01_WWR_Tristone.xlsx"
areas_path = r"c:\Users\X1\OneDrive\Documentos\Python_VS Code\GProA\GProA_EOSIS_Edge\docs\Documentos_EOSIS\Tristone_Area Breakdown_Calculator.xlsx"

print("Abriendo WBS Master...")
wb_master = openpyxl.load_workbook(wbs_path)

print("Copiando WWR...")
wb_wwr = openpyxl.load_workbook(wwr_path, data_only=True)
ws_wwr = wb_wwr["Hoja 3"]
ws_new_wwr = wb_master.create_sheet("Calculadora WWR")
copy_sheet(ws_wwr, ws_new_wwr)

print("Copiando Areas...")
wb_areas = openpyxl.load_workbook(areas_path, data_only=True)
ws_areas = wb_areas["Hoja 1"]
ws_new_areas = wb_master.create_sheet("Calculadora Areas")
copy_sheet(ws_areas, ws_new_areas)

print("Guardando Master...")
wb_master.save(wbs_path)
print("¡Fusión completada con openpyxl!")
